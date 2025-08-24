from __future__ import annotations

from io import BytesIO
from typing import Optional, List, Tuple

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert
from openpyxl import load_workbook

from app.db.database import get_db_session
from app.db.models import Lead
from app.schemas.leads import SMS, LeadBulkInsertResponse


router = APIRouter()


def _normalize_header(s: Optional[str]) -> str:
    return (s or "").strip().lower()


def _extract_rows_from_sheet(ws) -> List[Tuple[str, str, Optional[str]]]:
    # Expect headers: Lead Name, Lead Email, Lead Mobile (case-insensitive)
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {}
    for idx, name in enumerate(first_row, start=1):
        header_map[_normalize_header(name)] = idx

    required = ["lead name", "lead email"]
    for r in required:
        if r not in header_map:
            raise HTTPException(status_code=400, detail=f"Missing required column: {r}")

    name_idx = header_map.get("lead name")
    email_idx = header_map.get("lead email")
    mobile_idx = header_map.get("lead mobile")

    rows: List[Tuple[str, str, Optional[str]]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[name_idx - 1] if name_idx else None) or ""
        email = (row[email_idx - 1] if email_idx else None) or ""
        mobile = (row[mobile_idx - 1] if mobile_idx else None) if mobile_idx else None
        if not str(name).strip() and not str(email).strip():
            # skip completely empty lines
            continue
        rows.append((str(name).strip(), str(email).strip(), str(mobile).strip() if mobile is not None else None))
    return rows


async def _bulk_insert_leads(rows: List[Tuple[str, str, Optional[str]]], db: AsyncSession) -> Tuple[int, int, int, List[str]]:
    total = len(rows)
    if total == 0:
        return 0, 0, 0, []

    payload = [
        {"name": name, "email": email, "mobile": mobile}
        for name, email, mobile in rows
        if email  # ensure email is present
    ]

    valid = len(payload)
    if valid == 0:
        return total, 0, 0, ["No valid rows with email found"]

    stmt = insert(Lead).values(payload)
    stmt = stmt.on_conflict_do_nothing(index_elements=[Lead.email]).returning(Lead.id)

    result = await db.execute(stmt)
    inserted_ids = result.fetchall()
    inserted = len(inserted_ids)
    duplicate = valid - inserted
    failed = total - valid

    return total, inserted, duplicate, []


@router.post("/bulk-insert", response_model=LeadBulkInsertResponse, status_code=status.HTTP_201_CREATED)
async def bulk_insert_leads(
    file: Optional[UploadFile] = File(None, description="Excel file to upload (.xlsx)"),
    file_path: Optional[str] = Form(None, description="Local server path to Excel file (.xlsx)"),
    db: AsyncSession = Depends(get_db_session),
):
    if not file and not file_path:
        raise HTTPException(status_code=400, detail="Either file or file_path must be provided")

    try:
        if file is not None:
            if file.filename and not file.filename.lower().endswith(".xlsx"):
                raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
            content = await file.read()
            wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        else:
            if not str(file_path).lower().endswith(".xlsx"):
                raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        ws = wb.active
        rows = _extract_rows_from_sheet(ws)
        total, inserted, duplicate, errors = await _bulk_insert_leads(rows, db)
        return LeadBulkInsertResponse(
            total_rows=total,
            inserted_count=inserted,
            duplicate_count=duplicate,
            failed_count=total - (inserted + duplicate),
            errors=errors,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {e}")


@router.post("/sms", status_code=status.HTTP_200_OK)
async def sms(
    sms: SMS,
    db: AsyncSession = Depends(get_db_session),
):
    """Send personalized SMS to all leads using SirenAgentToolkit."""
    try:
        from sqlalchemy import select
        result = await db.execute(
            select(Lead.name, Lead.mobile).where(Lead.mobile.isnot(None))
        )
        leads = result.fetchall()
        
        if not leads:
            raise HTTPException(status_code=404, detail="No leads with mobile numbers found")
        
        from app.core.config import settings
        from langchain_openai import ChatOpenAI
        from langchain.agents import AgentExecutor, create_tool_calling_agent
        from langchain_core.prompts import ChatPromptTemplate
        from agenttoolkit.langchain import SirenAgentToolkit
        
        llm = ChatOpenAI(
            model="gpt-4o-mini",
            temperature=0.7,
            api_key=settings.OPENAI_API_KEY
        )
        
        siren_toolkit = SirenAgentToolkit(
            api_key=settings.SIREN_API_KEY,
            configuration={
                "actions": {
                    "messaging": {
                        "create": True,
                        "read": True,
                    },
                    "templates": {
                        "read": True,
                        "create": True,
                    }
                },
            },
        )
        
        prompt = ChatPromptTemplate.from_messages([
            ("system", """You are a helpful assistant that can send SMS messages using Siren. 
            
            IMPORTANT: For SMS channel, always provide:
            - channel: "SMS"
            - recipient_value: the mobile number (e.g., "+919188065817")
            - body: the message content
            
            When sending messages, personalize them with the recipient's name and make them conversational and friendly. Keep SMS messages under 160 characters when possible."""),
            ("human", "{input}"),
            ("placeholder", "{agent_scratchpad}"),
        ])
        
        tools = siren_toolkit.get_tools()
        
        
        agent = create_tool_calling_agent(llm, tools, prompt)
        agent_executor = AgentExecutor(agent=agent, tools=tools, verbose=True)
        
        sent_count = 0
        failed_count = 0
        results = []
        
        for name, mobile in leads:
            try:
                agent_input = f"""
                Send a message using the following details:
                - Channel: SMS
                - Recipient mobile number: {mobile}
                - Message content: Personalize this for "{name}": {sms.content}
                
                Make the message personal by adding their name naturally and keep it conversational and friendly.
                Ensure you provide the recipient mobile number in the correct field for SMS channel.
                """
                
                result = agent_executor.invoke({"input": agent_input})
                
                sent_count += 1
                results.append({
                    "name": name,
                    "mobile": mobile, 
                    "status": "sent",
                    "agent_response": result.get("output", "Message sent successfully")
                })
                    
            except Exception as e:
                failed_count += 1
                results.append({
                    "name": name,
                    "mobile": mobile,
                    "status": "failed", 
                    "error": str(e)
                })
        
        return {
            "total_leads": len(leads),
            "sent_count": sent_count,
            "failed_count": failed_count,
            "results": results,
            "message": f"SMS broadcast completed: {sent_count} sent, {failed_count} failed"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SMS broadcast failed: {str(e)}")


@router.post("/linkedin-product-updates", status_code=status.HTTP_200_OK)
async def linkedin_product_updates(
    product_info: dict,
    db: AsyncSession = Depends(get_db_session),
):
    """Crawl all leads' LinkedIn profiles, analyze their posts, and send personalized product updates."""
    try:
        from sqlalchemy import select
        result = await db.execute(
            select(Lead.name, Lead.email).where(Lead.email.isnot(None))
        )
        leads = result.fetchall()
        
        if not leads:
            raise HTTPException(status_code=404, detail="No leads found")
            
        from app.core.config import settings
        from langchain_openai import ChatOpenAI
        from langchain.agents import AgentExecutor, create_tool_calling_agent
        from langchain_core.prompts import ChatPromptTemplate
        from app.core.graphs.tools.linkedin.tool_registry import get_linkedin_tools
        from agenttoolkit.langchain import SirenAgentToolkit
        
        # Set up LLM
        llm = ChatOpenAI(
            model="gpt-4o-mini",
            temperature=0.7,
            api_key=settings.OPENAI_API_KEY
        )
        
        # Get LinkedIn tools for profile/post analysis
        linkedin_tools = get_linkedin_tools()
        
        # Set up Siren toolkit for sending updates
        siren_toolkit = SirenAgentToolkit(
            api_key=settings.SIREN_API_KEY,
            configuration={
                "actions": {
                    "messaging": {"create": True, "read": True},
                    "templates": {"read": True, "create": True, "update": True, "delete": True},
                    "users": {"create": True, "update": True, "delete": True, "read": True},
                    "workflows": {"trigger": True, "schedule": True},
                },
            },
        )
        
        # Combine LinkedIn and Siren tools
        all_tools = linkedin_tools + siren_toolkit.get_tools()
        
        # Create comprehensive prompt
        prompt = ChatPromptTemplate.from_messages([
            ("system", f"""You are an AI assistant that can analyze LinkedIn profiles and send personalized product updates.
            
            TASK: For each lead, you should:
            1. Search for their LinkedIn profile using their name and email
            2. Get their recent posts and activities
            3. Analyze their interests and professional focus
            4. Create a personalized email about our new product that resonates with their interests
            5. Send the email via Siren API
            
            PRODUCT INFO: {product_info}
            
            For email sending, use:
            - channel: "EMAIL"
            - recipient_value: the email address
            - subject: personalized subject line
            - body: personalized product update message
            
            Make each message highly relevant to their LinkedIn activity and interests."""),
            ("human", "{input}"),
            ("placeholder", "{agent_scratchpad}"),
        ])
        
        # Create agent
        agent = create_tool_calling_agent(llm, all_tools, prompt)
        agent_executor = AgentExecutor(agent=agent, tools=all_tools, verbose=True)
        
        processed_count = 0
        sent_count = 0
        failed_count = 0
        results = []
        
        # Process each lead
        for name, email in leads:
            try:
                processed_count += 1
                
                agent_input = f"""
                Process lead: {name} (Email: {email})
                
                Steps:
                1. Search LinkedIn for "{name}" using their email domain for context
                2. Get their profile information and recent posts/activities
                3. Analyze their professional interests and recent engagement
                4. Create a personalized email about our product that connects with their interests
                5. Send the personalized email to {email}
                
                Focus on making the product update relevant to their LinkedIn activity.
                """
                
                result = agent_executor.invoke({"input": agent_input})
                
                sent_count += 1
                results.append({
                    "name": name,
                    "email": email,
                    "status": "processed",
                    "agent_response": result.get("output", "LinkedIn analysis and email sent successfully")
                })
                
            except Exception as e:
                failed_count += 1
                results.append({
                    "name": name,
                    "email": email,
                    "status": "failed",
                    "error": str(e)
                })
        
        return {
            "total_leads": len(leads),
            "processed_count": processed_count,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "product_info": product_info,
            "results": results,
            "message": f"LinkedIn product updates completed: {sent_count} sent, {failed_count} failed"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LinkedIn product updates failed: {str(e)}")